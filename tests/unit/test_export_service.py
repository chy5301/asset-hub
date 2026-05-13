"""ExportService 单元测试."""

from __future__ import annotations

import re
from datetime import date
from io import BytesIO

import openpyxl
from sqlmodel import Session

from asset_hub.models.asset import AssetStatus
from asset_hub.services.asset import AssetService
from asset_hub.services.asset_type import TypeService
from asset_hub.services.export import STATUS_HEX, STATUS_LABELS, ExportService


class TestStatusDicts:
    def test_status_labels_covers_6_enum_values(self):
        from asset_hub.models.asset import AssetStatus
        from asset_hub.services.export import STATUS_LABELS

        # 应覆盖全部 6 个 enum 值
        assert set(STATUS_LABELS.keys()) == set(AssetStatus)
        assert len(STATUS_LABELS) == 6

    def test_status_labels_chinese(self):
        assert STATUS_LABELS[AssetStatus.IN_USE] == "在用"
        assert STATUS_LABELS[AssetStatus.IDLE] == "闲置"
        assert STATUS_LABELS[AssetStatus.MAINTENANCE] == "送修"
        assert STATUS_LABELS[AssetStatus.BROKEN] == "故障"
        assert STATUS_LABELS[AssetStatus.RETIRED] == "退役"
        assert STATUS_LABELS[AssetStatus.DISPOSED] == "注销"

    def test_status_hex_covers_6_enum_values(self):
        from asset_hub.models.asset import AssetStatus
        from asset_hub.services.export import STATUS_HEX

        assert set(STATUS_HEX.keys()) == set(AssetStatus)
        assert len(STATUS_HEX) == 6

    def test_status_hex_format(self):
        # ARGB hex: 8 大写 char, 前 2 char 是 FF (full alpha)
        for hex_val in STATUS_HEX.values():
            assert re.fullmatch(r"FF[0-9A-F]{6}", hex_val), (
                f"非法 ARGB hex: {hex_val!r}, 期望形如 'FFRRGGBB'"
            )


class TestResolveCustomFields:
    def test_returns_empty_when_type_id_none(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)
        assert svc._resolve_custom_fields(None) == []

    def test_returns_type_custom_fields_when_locked(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(
            name="Laptop",
            code_prefix="NB",
            custom_fields=[
                {"key": "sn", "label": "铭牌编号", "type": "string", "required": False},
                {"key": "cpu", "label": "CPU", "type": "string", "required": False},
            ],
        )

        fields = svc._resolve_custom_fields(t.id)
        assert len(fields) == 2
        assert fields[0].key == "sn"
        assert fields[0].label == "铭牌编号"
        assert fields[1].key == "cpu"


class TestBuildRows:
    def test_column_order_no_custom_fields(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="Laptop", code_prefix="NB", custom_fields=[])
        a = asset_svc.register(name="A1", type_id=t.id, custom_data={})

        rows = svc._build_rows([a], custom_fields=[])
        assert len(rows) == 1
        keys = list(rows[0].keys())
        assert keys == [
            "资产编号",
            "名称",
            "型号",
            "类型",
            "状态",
            "保管人",
            "位置",
            "闲置天数",
            "入账日期",
            "铭牌编号",
            "备注",
        ]

    def test_status_chinese_label(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="LA", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})

        rows = svc._build_rows([a], custom_fields=[])
        assert rows[0]["状态"] == "闲置"  # register 默认 IDLE

    def test_acquired_at_iso_date(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="LB", custom_fields=[])
        a = asset_svc.register(
            name="X", type_id=t.id, custom_data={}, acquired_at=date(2026, 1, 15)
        )

        rows = svc._build_rows([a], custom_fields=[])
        assert rows[0]["入账日期"] == "2026-01-15"

    def test_acquired_at_none_empty_string(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="LC", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})

        rows = svc._build_rows([a], custom_fields=[])
        assert rows[0]["入账日期"] == ""

    def test_holder_location_sn_notes_none_empty_string(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="LD", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})

        rows = svc._build_rows([a], custom_fields=[])
        assert rows[0]["保管人"] == ""
        assert rows[0]["位置"] == ""
        assert rows[0]["铭牌编号"] == ""
        assert rows[0]["备注"] == ""

    def test_custom_fields_flattened_when_provided(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(
            name="Laptop",
            code_prefix="NBX",
            custom_fields=[
                {
                    "key": "sn",
                    "label": "铭牌编号 (custom)",
                    "type": "string",
                    "required": False,
                },
                {"key": "cpu", "label": "CPU", "type": "string", "required": False},
            ],
        )
        a = asset_svc.register(
            name="X",
            type_id=t.id,
            custom_data={"sn": "SN-001", "cpu": "i9-12900K"},
        )

        custom_fields = svc._resolve_custom_fields(t.id)
        rows = svc._build_rows([a], custom_fields=custom_fields)

        keys = list(rows[0].keys())
        assert keys[-2:] == ["铭牌编号 (custom)", "CPU"]
        assert rows[0]["铭牌编号 (custom)"] == "SN-001"
        assert rows[0]["CPU"] == "i9-12900K"

    def test_custom_field_missing_value_empty_string(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(
            name="L",
            code_prefix="LE",
            custom_fields=[
                {"key": "sn", "label": "SN", "type": "string", "required": False},
            ],
        )
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})

        rows = svc._build_rows([a], custom_fields=svc._resolve_custom_fields(t.id))
        assert rows[0]["SN"] == ""

    def test_idle_days_after_annotate(self, session: Session):
        """idle_days 由 annotate_idle_days 注入; _build_rows 期望 caller 已 annotate."""
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="LF", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})  # IDLE
        annotated = asset_svc.annotate_idle_days([a])  # 注入 _idle_days_value

        rows = svc._build_rows(annotated, custom_fields=[])
        assert rows[0]["闲置天数"] in {"0", "1"}  # 刚 register, 0 或 1 天

    def test_idle_days_none_without_annotate(self, session: Session):
        """未 annotate 的 asset, idle_days property 返 None → '' 兜底."""
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="LG", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})  # 不调 annotate

        rows = svc._build_rows([a], custom_fields=[])
        assert rows[0]["闲置天数"] == ""

    def test_empty_assets_returns_empty_list(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        rows = svc._build_rows([], custom_fields=[])
        assert rows == []


class TestRenderCsv:
    def test_starts_with_utf8_bom(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="CA", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_csv(rows)
        assert data.startswith(b"\xef\xbb\xbf"), f"missing BOM, got {data[:10]!r}"

    def test_header_row_chinese(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="CB", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_csv(rows)
        text = data.decode("utf-8-sig")
        first_line = text.splitlines()[0]
        assert first_line.startswith("资产编号,名称,型号,类型,状态,")

    def test_empty_rows_writes_header_only(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        column_names = [
            "资产编号",
            "名称",
            "型号",
            "类型",
            "状态",
            "保管人",
            "位置",
            "闲置天数",
            "入账日期",
            "铭牌编号",
            "备注",
        ]
        data = svc._render_csv([], column_names=column_names)
        text = data.decode("utf-8-sig")
        lines = text.strip().splitlines()
        assert len(lines) == 1
        assert lines[0] == ",".join(column_names)

    def test_csv_escape_comma_quote_newline(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="CC", custom_fields=[])
        a = asset_svc.register(
            name='含,逗号"引号\n换行',
            type_id=t.id,
            custom_data={},
        )
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_csv(rows, column_names=list(rows[0].keys()))
        text = data.decode("utf-8-sig")
        # 含逗号 / 引号 / 换行的字段必须包在双引号内, 内部双引号 escape 为 ""
        assert '"含,逗号""引号' in text


class TestRenderXlsx:
    @staticmethod
    def _load(data: bytes):
        return openpyxl.load_workbook(BytesIO(data))

    def test_sheet_name_assets_list(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="XA", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_xlsx(rows, column_names=list(rows[0].keys()))
        wb = self._load(data)
        assert "资产清单" in wb.sheetnames

    def test_header_row_bold(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="XB", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_xlsx(rows, column_names=list(rows[0].keys()))
        ws = self._load(data)["资产清单"]
        assert ws.cell(row=1, column=1).value == "资产编号"
        assert ws.cell(row=1, column=1).font.bold is True

    def test_freeze_panes_a2(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="XC", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_xlsx(rows, column_names=list(rows[0].keys()))
        ws = self._load(data)["资产清单"]
        assert ws.freeze_panes == "A2"

    def test_autofilter_full_range(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="XD", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_xlsx(rows, column_names=list(rows[0].keys()))
        ws = self._load(data)["资产清单"]
        # 11 列 (固定，v2.0 PR-3 加 "型号") + 1 row header + 1 row data = "A1:K2"
        assert ws.auto_filter.ref == "A1:K2"

    def test_status_cell_filled_with_status_hex(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="XE", custom_fields=[])
        a = asset_svc.register(name="X", type_id=t.id, custom_data={})  # IDLE
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_xlsx(rows, column_names=list(rows[0].keys()))
        ws = self._load(data)["资产清单"]
        # 状态列因 "型号" 列插入而后移到第 5 列 (E)
        status_cell = ws.cell(row=2, column=5)
        assert status_cell.value == "闲置"
        # PatternFill 验 fgColor.rgb 与 STATUS_HEX[IDLE] 一致
        assert status_cell.fill.fgColor.rgb == STATUS_HEX[AssetStatus.IDLE]

    def test_empty_rows_only_header(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        column_names = [
            "资产编号",
            "名称",
            "型号",
            "类型",
            "状态",
            "保管人",
            "位置",
            "闲置天数",
            "入账日期",
            "铭牌编号",
            "备注",
        ]
        data = svc._render_xlsx([], column_names=column_names)
        ws = self._load(data)["资产清单"]
        assert ws.cell(row=1, column=1).value == "资产编号"
        assert ws.cell(row=2, column=1).value is None

    def test_notes_column_width_capped(self, session: Session):
        """spec §B.3: notes 列宽 cap 60, 其他列 cap 50."""
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="XF", custom_fields=[])
        a = asset_svc.register(
            name="X",
            type_id=t.id,
            custom_data={},
            notes="x" * 100,  # 100 char 长文本, 列宽应 cap 60
        )
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_xlsx(rows, column_names=list(rows[0].keys()))
        ws = self._load(data)["资产清单"]
        # 备注是固定第 10 列 (J)
        assert ws.column_dimensions["J"].width is not None
        assert ws.column_dimensions["J"].width <= 60

    def test_wrap_text_enabled_on_data_cells(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="XG", custom_fields=[])
        a = asset_svc.register(
            name="X", type_id=t.id, custom_data={}, notes="long text"
        )
        rows = svc._build_rows([a], custom_fields=[])

        data = svc._render_xlsx(rows, column_names=list(rows[0].keys()))
        ws = self._load(data)["资产清单"]
        notes_cell = ws.cell(row=2, column=10)
        assert notes_cell.alignment.wrap_text is True


class TestExport:
    def test_format_csv_returns_bom_bytes(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="EA", custom_fields=[])
        asset_svc.register(name="X", type_id=t.id, custom_data={})

        data, filename = svc.export(format="csv")
        assert data.startswith(b"\xef\xbb\xbf")
        assert filename.endswith(".csv")

    def test_format_xlsx_returns_pk_magic(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(name="L", code_prefix="EB", custom_fields=[])
        asset_svc.register(name="X", type_id=t.id, custom_data={})

        data, filename = svc.export(format="xlsx")
        # XLSX 是 zip, magic bytes "PK\x03\x04"
        assert data.startswith(b"PK")
        assert filename.endswith(".xlsx")

    def test_filename_format(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        _, filename = svc.export(format="csv")
        # spec §B.9: assets-YYYYMMDD-HHMM.csv
        assert re.fullmatch(r"assets-\d{8}-\d{4}\.csv", filename), (
            f"unexpected filename: {filename!r}"
        )

    def test_filter_passed_through_with_type_id_lock(self, session: Session):
        """type_id 锁定 → 仅该 type 的 assets + custom_fields 平铺."""
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t1 = type_svc.create_type(
            name="Laptop",
            code_prefix="EC",
            custom_fields=[
                {"key": "sn", "label": "SN", "type": "string", "required": False},
            ],
        )
        t2 = type_svc.create_type(name="GPU", code_prefix="ED", custom_fields=[])
        asset_svc.register(name="A1", type_id=t1.id, custom_data={"sn": "NB-001"})
        asset_svc.register(name="A2", type_id=t1.id, custom_data={"sn": "NB-002"})
        asset_svc.register(name="A3", type_id=t2.id, custom_data={})  # GPU

        data, _ = svc.export(format="csv", type_id=t1.id)
        text = data.decode("utf-8-sig")
        non_empty_lines = [line for line in text.splitlines() if line.strip()]
        # 1 header + 2 t1 资产 = 3 行
        assert len(non_empty_lines) == 3
        # 含 SN 平铺列
        assert "SN" in non_empty_lines[0]
        # 不含 GPU asset
        assert "A3" not in text

    def test_filter_no_type_id_no_custom_flatten(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        t = type_svc.create_type(
            name="Laptop",
            code_prefix="EE",
            custom_fields=[
                {"key": "sn", "label": "SN", "type": "string", "required": False},
            ],
        )
        asset_svc.register(name="A1", type_id=t.id, custom_data={"sn": "NB-001"})

        data, _ = svc.export(format="csv")  # 不传 type_id
        text = data.decode("utf-8-sig")
        first_line = text.splitlines()[0]
        # SN 平铺列不应出现
        assert "SN" not in first_line

    def test_zero_results_csv_only_header(self, session: Session):
        type_svc = TypeService(session)
        asset_svc = AssetService(session)
        svc = ExportService(session, asset_svc, type_svc)

        # 完全空 db
        data, _ = svc.export(format="csv")
        text = data.decode("utf-8-sig")
        non_empty_lines = [line for line in text.splitlines() if line.strip()]
        assert len(non_empty_lines) == 1  # 仅 header
        assert non_empty_lines[0].startswith("资产编号,")
