"""
导出服务: CSV/XLSX 资产清单导出.

spec: docs/superpowers/specs/2026-05-07-m3c-export-design.md
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session

from asset_hub.api.schemas.asset_type import CustomFieldDef
from asset_hub.models.asset import Asset, AssetStatus
from asset_hub.services.asset import AssetService
from asset_hub.services.asset_type import TypeService

# spec §B.4: 状态字段写人类标签 (与 frontend STATUS_META.label 同义)
# v2.0：6 态全两字对齐
STATUS_LABELS: dict[AssetStatus, str] = {
    AssetStatus.IDLE: "闲置",
    AssetStatus.IN_USE: "在用",
    AssetStatus.MAINTENANCE: "送修",
    AssetStatus.BROKEN: "故障",  # v2.0 新
    AssetStatus.RETIRED: "退役",
    AssetStatus.DISPOSED: "注销",  # v2.0 改名
}

# 反向: label → enum (用于 _render_xlsx 染色时把 row 中的中文 status 反查回 enum)
_LABEL_TO_STATUS: dict[str, AssetStatus] = {v: k for k, v in STATUS_LABELS.items()}

# spec §B.7: 6 态 light 模式 OKLCH 转 sRGB ARGB hex.
# v2.0 加 BROKEN hex（基于 globals.css line ~121: oklch(0.93 0.13 30) → 红橙色调）
# 与 frontend 视觉对齐, 但导出文件永远用 light hex (打印友好, 与浏览器
# dark/light 切换无关).
STATUS_HEX: dict[AssetStatus, str] = {
    AssetStatus.IDLE: "FFE4ECF5",
    AssetStatus.IN_USE: "FFBBF5CD",
    AssetStatus.MAINTENANCE: "FFFFDCA8",
    AssetStatus.BROKEN: "FFFCCFC1",  # v2.0 新（oklch(0.93 0.13 30) 红橙近似）
    AssetStatus.RETIRED: "FFE5E8EB",
    AssetStatus.DISPOSED: "FFEEEEEE",
}


class ExportService:
    """资产清单 CSV/XLSX 导出. spec §B.2 / §2.2."""

    _SHEET_NAME = "资产清单"
    _STATUS_COLUMN_HEADER = "状态"
    _NOTES_COLUMN_HEADER = "备注"
    _COL_WIDTH_DEFAULT_CAP = 50
    _COL_WIDTH_NOTES_CAP = 60

    # spec §B.3: 10 固定列, 顺序严格 (custom fields 平铺接在尾)
    _FIXED_COLUMN_NAMES: list[str] = [
        "资产编号", "名称", "类型", "状态", "保管人", "位置",
        "闲置天数", "入账日期", "铭牌编号", "备注",
    ]

    def __init__(
        self,
        session: Session,
        asset_service: AssetService,
        type_service: TypeService,
    ) -> None:
        self.session = session
        self.asset_service = asset_service
        self.type_service = type_service

    def _resolve_custom_fields(
        self, type_id: uuid.UUID | None
    ) -> list[CustomFieldDef]:
        """spec §B.2: type_id 显式锁定时返 type.custom_fields, 否则 []."""
        if type_id is None:
            return []
        return self.type_service.get_type(type_id).custom_fields

    def _build_rows(
        self,
        assets: list[Asset],
        custom_fields: list[CustomFieldDef],
    ) -> list[dict[str, str]]:
        """spec §B.3: 10 固定列 + custom_fields 平铺. 列序严格.

        idle_days 期望 caller 已调 AssetService.annotate_idle_days(assets);
        未 annotate 时 asset.idle_days 返 None → 兜底空字符串.
        """
        rows: list[dict[str, str]] = []
        for a in assets:
            row: dict[str, str] = {
                "资产编号": a.asset_code,
                "名称": a.name,
                "类型": a.type_name or "",
                "状态": STATUS_LABELS[a.status],
                "保管人": a.holder or "",
                "位置": a.location or "",
                "闲置天数": str(a.idle_days) if a.idle_days is not None else "",
                "入账日期": a.acquired_at.isoformat() if a.acquired_at else "",
                "铭牌编号": a.serial_number or "",
                "备注": a.notes or "",
            }
            for field in custom_fields:
                header = field.label or field.key
                row[header] = str((a.custom_data or {}).get(field.key, ""))
            rows.append(row)
        return rows

    def _render_csv(
        self,
        rows: list[dict[str, str]],
        column_names: list[str] | None = None,
    ) -> bytes:
        """spec §B.6: UTF-8 BOM + stdlib csv.writer.

        column_names 在 rows 非空时可省 (从 rows[0].keys() 推断);
        rows 空时必传 (0 结果仅 header 场景).
        """
        if column_names is None:
            if not rows:
                return b"\xef\xbb\xbf"
            column_names = list(rows[0].keys())

        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=column_names, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

        return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")

    def _render_xlsx(
        self,
        rows: list[dict[str, str]],
        column_names: list[str],
    ) -> bytes:
        """spec §B.7 / §B.7.1: openpyxl + 6 态 PatternFill + freeze A2 + autofilter."""
        wb = Workbook()
        ws = wb.active
        ws.title = self._SHEET_NAME

        # Header row 1, bold
        bold_font = Font(bold=True)
        for col_idx, name in enumerate(column_names, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = bold_font

        # Data rows
        wrap_align = Alignment(wrap_text=True, vertical="top")
        # 预生成 6 态 PatternFill 对象, 与 Alignment/Font 一样在 method 顶部复用
        status_fills: dict[AssetStatus, PatternFill] = {
            status: PatternFill(
                start_color=hex_val,
                end_color=hex_val,
                fill_type="solid",
            )
            for status, hex_val in STATUS_HEX.items()
        }
        status_col_idx = (
            column_names.index(self._STATUS_COLUMN_HEADER) + 1
            if self._STATUS_COLUMN_HEADER in column_names
            else None
        )
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, name in enumerate(column_names, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(name, ""))
                cell.alignment = wrap_align
                if col_idx == status_col_idx:
                    status_value = row.get(name, "")
                    status_enum = self._label_to_enum(status_value)
                    if status_enum is not None:
                        cell.fill = status_fills[status_enum]

        # Freeze + autofilter
        ws.freeze_panes = "A2"
        max_row = len(rows) + 1
        max_col_letter = get_column_letter(len(column_names))
        ws.auto_filter.ref = f"A1:{max_col_letter}{max_row}"

        # Column width auto-fit cap
        for col_idx, name in enumerate(column_names, start=1):
            letter = get_column_letter(col_idx)
            cap = (
                self._COL_WIDTH_NOTES_CAP
                if name == self._NOTES_COLUMN_HEADER
                else self._COL_WIDTH_DEFAULT_CAP
            )
            max_chars = len(name)
            for row in rows:
                value = row.get(name, "")
                # 中文字符按 ~2x 算 (XLSX 列宽单位约等于英文字符)
                width = sum(2 if ord(c) > 127 else 1 for c in value)
                if width > max_chars:
                    max_chars = width
            ws.column_dimensions[letter].width = min(max_chars + 2, cap)

        # 序列化 bytes
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _label_to_enum(label: str) -> AssetStatus | None:
        """状态中文标签反查 enum (仅用于 XLSX 染色; 找不到返 None 不染色)."""
        return _LABEL_TO_STATUS.get(label)

    def export(
        self,
        format: Literal["csv", "xlsx"],
        type_id: uuid.UUID | None = None,
        status: AssetStatus | None = None,
        holder: str | None = None,
        q: str | None = None,
        include_retired: bool = False,
        include_disposed: bool = False,
    ) -> tuple[bytes, str]:
        """spec §2.2: 整合 list_assets + annotate_idle_days + 渲染. 返 (bytes, filename).

        spec §B.10: 整 filter 集导出, 强制 sort_by=None / limit=None / offset=None,
        不分页不排序.
        """
        assets = self.asset_service.list_assets(
            type_id=type_id,
            status=status,
            holder=holder,
            q=q,
            include_retired=include_retired,
            include_disposed=include_disposed,
            sort_by=None,
            sort_order="desc",
            limit=None,
            offset=None,
        )
        # 关键: annotate idle_days 让 _build_rows "闲置天数" 列有值
        # (Asset.idle_days @property 仅在 _idle_days_value 注入后非 None)
        assets = self.asset_service.annotate_idle_days(assets)
        custom_fields = self._resolve_custom_fields(type_id)
        rows = self._build_rows(assets, custom_fields)

        column_names = list(self._FIXED_COLUMN_NAMES)
        for field in custom_fields:
            column_names.append(field.label or field.key)

        filename = self._build_filename(format)
        if format == "csv":
            return self._render_csv(rows, column_names=column_names), filename
        return self._render_xlsx(rows, column_names=column_names), filename

    @staticmethod
    def _build_filename(format: Literal["csv", "xlsx"]) -> str:
        """spec §B.9: assets-YYYYMMDD-HHMM.{csv,xlsx}."""
        stamp = datetime.now().strftime("%Y%m%d-%H%M")
        return f"assets-{stamp}.{format}"
